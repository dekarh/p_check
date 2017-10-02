# -*- coding: utf-8 -*-

import sys
import time
import datetime
import openpyxl
from openpyxl import Workbook
from mysql.connector import MySQLConnection
from lib import l, read_config

IN_SNILS = ['СНИЛС', 'СтраховойНомер', 'Страховой_номер', 'Страховой Номер', 'Номер СНИЛС']
IN_SERIA = ['Серия','серия','Серия_документа','Паспорт_серия', 'Серия паспорта']
IN_NUMBER = ['Номер','номер','Номер_документа','Паспорт_номер','Номер паспорта']


dbconfig = read_config(section='mysql')
dbconn = MySQLConnection(**dbconfig)  # Открываем БД из конфиг-файла


# print(sys.argv[1])

workbooks =  []
sheets = []
for i, xlsx_file in enumerate(sys.argv):                              # Загружаем все xlsx файлы
    if i == 0:
        continue
    workbooks.append(openpyxl.load_workbook(filename=xlsx_file, read_only=True))
    sheets.append(workbooks[i-1][workbooks[i-1].sheetnames[0]])

total_rows = 0
sheets_keys = []
for i, sheet in enumerate(sheets):                                    # Маркируем нужные столбцы
    total_rows += sheet.max_row
    keys = {}
    for j, row in enumerate(sheet.rows):
        if j > 0:
            break
        for k, cell in enumerate(row):                                # Проверяем, чтобы был СНИЛС
            if cell.value in IN_SNILS:
                keys[IN_SNILS[0]] = k
            if cell.value in IN_SERIA:
                keys[IN_SERIA[0]] = k
            if cell.value in IN_NUMBER:
                keys[IN_NUMBER[0]] = k
        if len(keys) < 3:
            print('В файле "' + sys.argv[i+1] + '" отсутствует колонка со СНИЛС, номером или серией паспорта')
            time.sleep(3)
            sys.exit()
    sheets_keys.append(keys)

print('\n'+ datetime.datetime.now().strftime("%H:%M:%S") +' Начинаем проверку \n')

wb = Workbook(write_only=True)
ws = wb.create_sheet('Лист1')
ws.append([IN_SNILS[0], IN_SERIA[0], IN_NUMBER[0], 'Проверка', 'ВнутренДубль'])  # добавляем первую строку xlsx
perc_rows = 0
all_good = True
all_not_doubles = True
for i, sheet in enumerate(sheets):
    for j, row in enumerate(sheet.rows):
        if j == 0:
            continue
        rez = 'OK'
        read_cursor = dbconn.cursor()
        read_cursor.execute('SELECT p_seria, p_number FROM passport_blacklist WHERE p_seria = %s AND p_number = %s',
                            (l(row[keys[IN_SERIA[0]]].value),l(row[keys[IN_NUMBER[0]]].value)))
        row_msg = read_cursor.fetchall()
        if len(row_msg) > 0:
            rez = 'плохой'
            all_good = False
        else:
            rez = 'ОК'
        double = 'нет'
        read_cursor = dbconn.cursor()
        read_cursor.execute('SELECT `number` FROM saturn_crm.clients WHERE `number`= %s',
                            (l(row[keys[IN_SNILS[0]]].value),))
        row_msg = read_cursor.fetchall()
        if len(row_msg) > 0:
            double = 'дубль'
            all_not_doubles = False
        else:
            double = 'нет'
        ws.append([row[keys[IN_SNILS[0]]].value, row[keys[IN_SERIA[0]]].value, row[keys[IN_NUMBER[0]]].value, rez, double])
        if int(j/total_rows*100) > perc_rows:
            perc_rows = int(j/total_rows*100)
            print(datetime.datetime.now().strftime("%H:%M:%S") + '  обработано ' + str(perc_rows) + '%')

if len(sys.argv) > 2:
    append = '+'
else:
    append = ''

wb.save(sys.argv[1][0:sys.argv[1].rfind('.xlsx')] + append + '_pasp' '.xlsx')

if all_good:
    print(datetime.datetime.now().strftime("%H:%M:%S") + ' Все паспорта хорошие')
else:
    print(datetime.datetime.now().strftime("%H:%M:%S") + ' Есть плохие паспорта')

if all_not_doubles:
    print(datetime.datetime.now().strftime("%H:%M:%S") + ' Внутренних дублей нет')
else:
    print(datetime.datetime.now().strftime("%H:%M:%S") + ' Есть внутренние дубли')

print('\n'+ datetime.datetime.now().strftime("%H:%M:%S") +' Проверка закончена \n')

